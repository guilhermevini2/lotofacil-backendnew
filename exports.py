"""
exports.py - Exportacao para Excel (6 abas), CSV, TXT e PDF.
"""

import csv
from datetime import datetime
from pathlib import Path

from config import (
    RELATORIO_DIR, JOGOS_DIR,
    EXCEL_FONT, EXCEL_HEADER, EXCEL_ALT, EXCEL_GREEN, EXCEL_YELLOW, EXCEL_ORANGE,
    JANELAS_TENDENCIA,
)


def _ts():
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def _argb(h):
    return "FF" + h.upper()


# -- TXT ----------------------------------------------------------------------

def exportar_txt(concursos, stats, ranking_rows, jogos, dezenas_18,
                 bt, validacao, txt_val, txt_bt,
                 pares_quentes=None, trios_quentes=None, analise_pools=None):
    caminho = RELATORIO_DIR / f"relatorio_{_ts()}.txt"
    L = []
    L += [
        "=" * 65,
        "  LOTOFACIL PRO v2 - RELATORIO COMPLETO",
        f"  Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        "=" * 65,
        f"\n  Concursos analisados : {len(concursos)}",
        f"  Periodo              : {concursos[0]['data']} -> {concursos[-1]['data']}",
        "",
    ]

    L += ["", "  RANKING DAS DEZENAS", "-" * 65]
    L.append(f"  {'Pos':>3}  {'Dez':>3}  {'Score':>6}  {'Freq':>5}  {'Freq%':>6}  "
             f"{'Decay':>7}  {'Atraso':>6}  {'T10':>6}  {'T20':>6}  {'Pool':>5}")
    for r in ranking_rows:
        pool = "SIM" if r["dezena"] in set(dezenas_18) else ""
        L.append(
            f"  {r['posicao']:>3}  {r['dezena']:>3}  {r['score']:>6.4f}  "
            f"{r['freq_abs']:>5}  {r['freq_pct']:>5.1f}%  {r['freq_decay']:>7.2f}  "
            f"{r['atraso']:>6}  {r['tend_10']:>5.1f}%  {r['tend_20']:>5.1f}%  {pool:>5}"
        )

    L += ["", "  18 DEZENAS SELECIONADAS", "-" * 65]
    L.append("  " + "  ".join(f"{d:02d}" for d in dezenas_18))

    if pares_quentes:
        L += ["", "  TOP 20 PARES QUENTES", "-" * 65]
        for a, b, cnt in pares_quentes[:20]:
            pool_str = "(ambas no pool)" if a in set(dezenas_18) and b in set(dezenas_18) else ""
            L.append(f"  {a:02d} + {b:02d}  ->  {cnt} vezes juntos  {pool_str}")

    if trios_quentes:
        L += ["", "  TOP 10 TRIOS QUENTES", "-" * 65]
        for a, b, c, cnt in trios_quentes[:10]:
            L.append(f"  {a:02d} + {b:02d} + {c:02d}  ->  {cnt} vezes juntos")

    if analise_pools:
        L += ["", "  COMPARATIVO DE POOLS", "-" * 65]
        L.append(f"  {'Pool':>4}  {'Jogos':>5}  {'Custo':>8}  {'Cobertura':>10}")
        for p in analise_pools:
            L.append(f"  {p['pool_size']:>4}  {p['n_jogos']:>5}  "
                     f"R${p['custo']:>6.2f}  {p['prob_cobertura']:>9.2f}%")

    L += ["", txt_val, "", txt_bt]

    L += ["", "  JOGOS DO FECHAMENTO 18-15-14", "-" * 65]
    for i, jogo in enumerate(jogos, 1):
        L.append(f"  Jogo {i:02d}: " + "  ".join(f"{d:02d}" for d in jogo))

    L += [
        "",
        "  AVISO: Garantia de >=14 pontos SOMENTE se os 15 sorteados",
        "  estiverem dentro das 18 dezenas do pool acima.",
        "=" * 65,
    ]
    caminho.write_text("\n".join(L), encoding="utf-8")
    return caminho


# -- CSV ----------------------------------------------------------------------

def exportar_csv_ranking(ranking_rows):
    caminho = RELATORIO_DIR / f"ranking_{_ts()}.csv"
    if not ranking_rows:
        return caminho
    with caminho.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(ranking_rows[0].keys()))
        w.writeheader()
        w.writerows(ranking_rows)
    return caminho


def exportar_csv_jogos(jogos):
    caminho = JOGOS_DIR / f"jogos_{_ts()}.csv"
    with caminho.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Jogo"] + [f"D{i}" for i in range(1, 16)])
        for i, jogo in enumerate(jogos, 1):
            w.writerow([i] + jogo)
    return caminho


def exportar_csv_historico(concursos):
    caminho = RELATORIO_DIR / f"historico_{_ts()}.csv"
    with caminho.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Concurso", "Data"] + [f"D{i}" for i in range(1, 16)])
        for cs in concursos:
            w.writerow([cs["concurso"], cs["data"]] + cs["dezenas"])
    return caminho


def exportar_csv_pares(pares_quentes):
    caminho = RELATORIO_DIR / f"pares_quentes_{_ts()}.csv"
    with caminho.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Dezena_A", "Dezena_B", "Aparicoes_Juntas"])
        for a, b, cnt in pares_quentes:
            w.writerow([a, b, cnt])
    return caminho


# -- Excel --------------------------------------------------------------------

def exportar_excel(concursos, stats, ranking_rows, jogos, dezenas_18,
                   bt, pares_quentes=None, trios_quentes=None, analise_pools=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl nao instalado.")

    wb = Workbook()

    def hdr(cell, cor=EXCEL_HEADER):
        cell.font = Font(name=EXCEL_FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=_argb(cor))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def alt(cell, par=True, cor=None):
        if cor:
            cell.fill = PatternFill("solid", fgColor=_argb(cor))
        elif par:
            cell.fill = PatternFill("solid", fgColor=_argb(EXCEL_ALT))
        cell.font = Font(name=EXCEL_FONT, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def autowidth(ws, extra=2):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = mx + extra

    pool_set = set(dezenas_18)

    # Aba 1: Ranking
    ws1 = wb.active
    ws1.title = "Ranking"
    cabecalhos = (["Pos", "Dezena", "Score", "Freq", "Freq%", "Decay", "Atraso"]
                  + [f"Tend {n}" for n in JANELAS_TENDENCIA] + ["Pool"])
    ws1.row_dimensions[1].height = 28
    for c, h in enumerate(cabecalhos, 1):
        hdr(ws1.cell(row=1, column=c, value=h))
    for ri, row in enumerate(ranking_rows, 2):
        no_pool = row["dezena"] in pool_set
        vals = ([row["posicao"], row["dezena"], row["score"],
                 row["freq_abs"], row["freq_pct"], row["freq_decay"], row["atraso"]]
                + [row.get(f"tend_{n}", 0) for n in JANELAS_TENDENCIA]
                + ["SIM" if no_pool else ""])
        for ci, val in enumerate(vals, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            alt(cell, par=(ri % 2 == 0), cor=EXCEL_GREEN if no_pool else None)
    autowidth(ws1)

    # Aba 2: Jogos
    ws2 = wb.create_sheet("Jogos")
    for c, h in enumerate(["Jogo"] + [f"D{i}" for i in range(1, 16)] + ["Soma", "Pares"], 1):
        hdr(ws2.cell(row=1, column=c, value=h))
    for ji, jogo in enumerate(jogos, 2):
        ws2.cell(row=ji, column=1, value=ji - 1)
        for ci, d in enumerate(jogo, 2):
            alt(ws2.cell(row=ji, column=ci, value=d), par=(ji % 2 == 0))
        ws2.cell(row=ji, column=17, value=sum(jogo))
        ws2.cell(row=ji, column=18, value=sum(1 for d in jogo if d % 2 == 0))
    autowidth(ws2)

    # Aba 3: Historico
    ws3 = wb.create_sheet("Historico")
    for c, h in enumerate(["Concurso", "Data"] + [f"D{i}" for i in range(1, 16)], 1):
        hdr(ws3.cell(row=1, column=c, value=h))
    for ri, cs in enumerate(concursos, 2):
        ws3.cell(row=ri, column=1, value=cs["concurso"])
        ws3.cell(row=ri, column=2, value=cs["data"])
        for ci, d in enumerate(cs["dezenas"], 3):
            alt(ws3.cell(row=ri, column=ci, value=d), par=(ri % 2 == 0))
    autowidth(ws3)

    # Aba 4: Backtest
    ws4 = wb.create_sheet("Backtest")
    tipo = "DINAMICO" if bt.get("tipo") == "dinamico" else "FIXO"
    ws4.cell(row=1, column=1, value=f"Backtest {tipo}").font = Font(bold=True)
    ws4.cell(row=2, column=1, value=f"Concursos: {bt['total_concursos']}")
    ws4.cell(row=2, column=3, value=f"Dentro do pool: {bt['pct_dentro_pool']}%")
    for c, h in enumerate(["Concurso", "Data", "Pool Hits", "Dentro Pool", "Melhor Pts", "Melhor Jogo"], 1):
        hdr(ws4.cell(row=4, column=c, value=h))
    for ri, det in enumerate(bt["detalhes"], 5):
        vals = [det["concurso"], det["data"], det["pool_hits"],
                "Sim" if det["dentro_pool"] else "Nao",
                det["melhor_pts"], det.get("melhor_jogo", "-")]
        cor = EXCEL_GREEN if det["dentro_pool"] else None
        for ci, val in enumerate(vals, 1):
            alt(ws4.cell(row=ri, column=ci, value=val), par=(ri % 2 == 0), cor=cor)
    autowidth(ws4)

    # Aba 5: Pares Quentes
    if pares_quentes:
        ws5 = wb.create_sheet("Pares Quentes")
        for c, h in enumerate(["Pos", "Dezena A", "Dezena B", "Aparicoes", "No Pool"], 1):
            hdr(ws5.cell(row=1, column=c, value=h))
        for ri, (a, b, cnt) in enumerate(pares_quentes, 2):
            ambos_pool = a in pool_set and b in pool_set
            uma_pool   = (a in pool_set) != (b in pool_set)
            cor = EXCEL_GREEN if ambos_pool else (EXCEL_YELLOW if uma_pool else None)
            for ci, val in enumerate([ri - 1, a, b, cnt, "Ambos" if ambos_pool else ("1 delas" if uma_pool else "")], 1):
                alt(ws5.cell(row=ri, column=ci, value=val), par=(ri % 2 == 0), cor=cor)
        autowidth(ws5)

    # Aba 6: Pools Comparativo
    if analise_pools:
        ws6 = wb.create_sheet("Pools")
        for c, h in enumerate(["Pool", "Jogos", "Custo (R$)", "Cobertura (%)"], 1):
            hdr(ws6.cell(row=1, column=c, value=h))
        for ri, p in enumerate(analise_pools, 2):
            for ci, val in enumerate([p["pool_size"], p["n_jogos"], p["custo"], p["prob_cobertura"]], 1):
                alt(ws6.cell(row=ri, column=ci, value=val), par=(ri % 2 == 0))
        autowidth(ws6)

    caminho = RELATORIO_DIR / f"relatorio_{_ts()}.xlsx"
    wb.save(str(caminho))
    return caminho


# -- PDF ----------------------------------------------------------------------

def exportar_pdf(_, caminho_txt):
    caminho_pdf = Path(str(caminho_txt).replace(".txt", ".pdf"))
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm

        c = canvas.Canvas(str(caminho_pdf), pagesize=A4)
        w, h = A4
        margin, y = 2 * cm, h - 2 * cm
        c.setFont("Courier", 8)

        for linha in Path(caminho_txt).read_text(encoding="utf-8").splitlines():
            if y < 2 * cm:
                c.showPage()
                y = h - 2 * cm
                c.setFont("Courier", 8)
            c.drawString(margin, y, linha[:120])
            y -= 11

        c.save()
        return caminho_pdf
    except ImportError:
        print("  INFO: reportlab nao instalado - PDF ignorado. Execute: pip install reportlab")
        return caminho_txt
